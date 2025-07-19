import asyncio
import json
from groq import Groq
from openai import AsyncOpenAI
import ollama
from playwright.async_api import async_playwright
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import re
import uuid
import logging
import os

# Configure logging
log_dir = "logs"
if not os.path.exists(log_dir):
    os.makedirs(log_dir)

log_file = os.path.join(log_dir, f"query_processor_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)
logger.info("Logging initialized at %s", datetime.now().strftime('%Y-%m-%d %H:%M:%S %Z'))

class QueryProcessor:
    def __init__(self, llm_provider="groq", groq_api_key=None, openai_api_key=None, ollama_model="gemma3:1b"):
        """
        Initialize QueryProcessor with specified LLM provider.
        Args:
            llm_provider (str): 'groq', 'openai', or 'ollama'
            groq_api_key (str): Groq API key (optional if set in env)
            openai_api_key (str): OpenAI API key (optional if set in env)
            ollama_model (str): Ollama model name (default: 'llama3')
        """
        self.supported_sites = {
            'amazon': 'https://www.amazon.com',
            'flipkart': 'https://www.flipkart.com',
        }
        self.llm_provider = llm_provider.lower()
        
        # Initialize LLM client based on provider
        if self.llm_provider == "groq":
            self.api_key = groq_api_key or os.getenv("GROQ_API_KEY", None)
            if not self.api_key:
                logger.error("No Groq API key provided")
                raise ValueError("Groq API key is required")
            self.client = Groq(api_key=self.api_key)
            self.model = "mixtral-8x7b-32768"
            logger.info("Initialized Groq client with API key ending in: %s", self.api_key[-4:])
        elif self.llm_provider == "openai":
            self.api_key = openai_api_key or os.getenv("OPENAI_API_KEY", None)
            if not self.api_key:
                logger.error("No OpenAI API key provided")
                raise ValueError("OpenAI API key is required")
            self.client = AsyncOpenAI(api_key=self.api_key)
            self.model = "gpt-4o-mini"
            logger.info("Initialized OpenAI client with API key ending in: %s", self.api_key[-4:])
        elif self.llm_provider == "ollama":
            self.model = ollama_model
            self.client = ollama
            logger.info("Initialized Ollama client with model: %s", self.model)
            # Verify Ollama server is running
            try:
                ollama.list()
            except Exception as e:
                logger.error("Ollama server not running or model %s not available: %s", self.model, str(e))
                raise ValueError(f"Ollama server not running or model {self.model} not available")
        else:
            logger.error("Unsupported LLM provider: %s", self.llm_provider)
            raise ValueError(f"Unsupported LLM provider: {self.llm_provider}")
        
        logger.info("QueryProcessor initialized with provider: %s, supported sites: %s", self.llm_provider, self.supported_sites)

    async def parse_query(self, query):
        """Parse natural language query using the specified LLM provider"""
        logger.info("Parsing query with %s: %s", self.llm_provider, query)
        retries = 3
        for attempt in range(retries):
            try:
                prompt = f"""
                Analyze this user query: "{query}"
                Return a JSON object with:
                - query_type (product_search, price_comparison, flight_search)
                - target_websites (list of sites)
                - search_params (dict with category, budget, specific_product, etc.)
                Ensure the response is a valid JSON string.
                Example response:
                {{
                    "query_type": "product_search",
                    "target_websites": ["amazon", "flipkart"],
                    "search_params": {{"category": "trimmers", "budget": "₹1000", "specific_product": null}}
                }}
                """
                
                if self.llm_provider == "groq":
                    response = self.client.chat.completions.create(
                        model=self.model,
                        messages=[{"role": "user", "content": prompt}],
                        response_format={"type": "json_object"}
                    )
                    parsed_result = json.loads(response.choices[0].message.content)
                elif self.llm_provider == "openai":
                    response = await self.client.chat.completions.create(
                        model=self.model,
                        messages=[{"role": "user", "content": prompt}],
                        response_format={"type": "json_object"}
                    )
                    parsed_result = json.loads(response.choices[0].message.content)
                elif self.llm_provider == "ollama":
                    response = ollama.chat(
                        model=self.model,
                        messages=[{"role": "user", "content": prompt}],
                        options={"format": "json"}  # Enforce JSON output
                    )
                    content = response.get("message", {}).get("content", "")
                    if not content:
                        logger.warning("Empty response from Ollama for query: %s", query)
                        raise ValueError("Empty response from Ollama")
                    try:
                        parsed_result = json.loads(content)
                    except json.JSONDecodeError as e:
                        logger.error("Invalid JSON from Ollama: %s", content)
                        raise ValueError(f"Invalid JSON from Ollama: {str(e)}")
                
                logger.info("Query parsed successfully: %s", parsed_result)
                return parsed_result
            except Exception as e:
                logger.error("Attempt %d/%d failed for query '%s' with %s: %s", 
                            attempt + 1, retries, query, self.llm_provider, str(e), exc_info=True)
                if attempt < retries - 1:
                    await asyncio.sleep(2 ** attempt)  # Exponential backoff
                else:
                    raise
        logger.error("All attempts to parse query '%s' failed", query)
        raise Exception("Failed to parse query after retries")

    async def generate_mcp_workflow(self, parsed_query):
        """Generate MCP browser automation steps based on parsed query"""
        logger.info("Generating MCP workflow for parsed query: %s", parsed_query)
        try:
            workflow = []
            target_websites = parsed_query.get('target_websites', [])
            search_params = parsed_query.get('search_params', {})
            search_term = search_params.get('specific_product', '') or search_params.get('category', '')
            budget = search_params.get('budget', '')

            for site in target_websites:
                if site.lower() in self.supported_sites:
                    steps = [
                        {"command": "browser_navigate", "value": f"{self.supported_sites[site.lower()]}/s?k={search_term}"},
                        {"command": "browser_scroll", "value": 2000},
                        {"command": "browser_wait", "value": 2},
                    ]
                    if budget and site.lower() == 'flipkart':
                        steps.append({"command": "browser_type", "selector": "input[name='q']", "value": f"{search_term} under {budget}"})
                        steps.append({"command": "browser_press", "value": "Enter"})
                    workflow.append({"site": site, "steps": steps})
                    logger.debug("Generated steps for %s: %s", site, steps)
                else:
                    logger.warning("Unsupported site: %s", site)
            logger.info("MCP workflow generated with %d site(s)", len(workflow))
            return workflow
        except Exception as e:
            logger.error("Error generating MCP workflow: %s", str(e), exc_info=True)
            raise

    async def scrape_site(self, page, site, steps):
        """Execute MCP workflow and scrape data"""
        logger.info("Scraping site: %s", site)
        results = []
        
        for step in steps:
            try:
                logger.debug("Executing step: %s", step)
                if step['command'] == 'browser_navigate':
                    await page.goto(step['value'], wait_until='domcontentloaded')
                    logger.debug("Navigated to %s", step['value'])
                elif step['command'] == 'browser_scroll':
                    await page.evaluate(f"window.scrollBy(0, {step['value']})")
                    logger.debug("Scrolled by %d pixels", step['value'])
                elif step['command'] == 'browser_wait':
                    await page.wait_for_timeout(step['value'] * 1000)
                    logger.debug("Waited for %d seconds", step['value'])
                elif step['command'] == 'browser_type':
                    await page.fill(step['selector'], step['value'])
                    logger.debug("Typed '%s' into selector '%s'", step['value'], step['selector'])
                elif step['command'] == 'browser_press':
                    await page.keyboard.press(step['value'])
                    logger.debug("Pressed key: %s", step['value'])
            except Exception as e:
                logger.error("Error in MCP step for %s: %s", site, str(e), exc_info=True)

        # Site-specific scraping logic
        if site.lower() == 'amazon':
            try:
                items = await page.query_selector_all('.s-result-item')
                logger.info("Found %d items on Amazon", len(items))
                for item in items[:5]:
                    title_elem = await item.query_selector('h2')
                    price_elem = await item.query_selector('.a-price .a-offscreen')
                    title = await title_elem.inner_text() if title_elem else "N/A"
                    price = await price_elem.inner_text() if price_elem else "N/A"
                    results.append({
                        'site': 'Amazon',
                        'title': title.strip(),
                        'price': price.strip(),
                        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    })
                    logger.debug("Scraped item: %s, %s", title, price)
            except Exception as e:
                logger.error("Scraping error on Amazon: %s", str(e), exc_info=True)
        
        elif site.lower() == 'flipkart':
            try:
                await page.wait_for_selector('div._1AtVbE', timeout=10000)
                items = await page.query_selector_all('div._1AtVbE')
                logger.info("Found %d items on Flipkart", len(items))
                for item in items[:5]:
                    title_elem = await item.query_selector('a.s1Q9rs')
                    price_elem = await item.query_selector('div._30jeq3')
                    title = await title_elem.inner_text() if title_elem else "N/A"
                    price = await price_elem.inner_text() if price_elem else "N/A"
                    results.append({
                        'site': 'Flipkart',
                        'title': title.strip(),
                        'price': price.strip(),
                        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    })
                    logger.debug("Scraped item: %s, %s", title, price)
            except Exception as e:
                logger.error("Scraping error on Flipkart: %s", str(e), exc_info=True)

        logger.info("Scraped %d items from %s", len(results), site)
        return results

    def create_excel_report(self, data, query):
        """Generate Excel report with data, charts, and conditional formatting"""
        logger.info("Creating Excel report for query: %s", query)
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Query Results"

            # Headers
            headers = ['Site', 'Product Title', 'Price', 'Timestamp']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            logger.debug("Excel headers set: %s", headers)

            # Data
            df = pd.DataFrame(data)
            for row, record in enumerate(data, 2):
                ws.cell(row=row, column=1).value = record['site']
                ws.cell(row=row, column=2).value = record['title']
                ws.cell(row=row, column=3).value = record['price']
                ws.cell(row=row, column=4).value = record['timestamp']
            logger.debug("Wrote %d rows to Excel", len(data))

            # Conditional formatting: Highlight prices below average
            if not df.empty:
                prices = []
                for price in df['price']:
                    try:
                        numeric = float(re.sub(r'[^\d.]', '', price))
                        prices.append(numeric)
                    except:
                        prices.append(0)
                
                avg_price = sum(prices) / len(prices) if prices else 0
                logger.debug("Average price calculated: %s", avg_price)
                for row in range(2, len(data) + 2):
                    price_cell = ws.cell(row=row, column=3)
                    try:
                        price_value = float(re.sub(r'[^\d.]', '', price_cell.value or '0'))
                        if price_value < avg_price:
                            price_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    except:
                        logger.warning("Failed to apply conditional formatting for row %d", row)

            # Add filter
            ws.auto_filter.ref = ws.dimensions
            logger.debug("Auto-filter applied to Excel sheet")

            # Create bar chart
            if not df.empty:
                chart = BarChart()
                chart.title = "Price Comparison by Site"
                chart.x_axis.title = "Product"
                chart.y_axis.title = "Price"
                
                df['numeric_price'] = prices
                chart_data = Reference(ws, min_col=3, min_row=1, max_row=len(data)+1)
                chart_cats = Reference(ws, min_col=2, min_row=2, max_row=len(data)+1)
                chart.add_data(chart_data, titles_from_data=True)
                chart.set_categories(chart_cats)
                ws.add_chart(chart, "F5")
                logger.debug("Bar chart added to Excel")

            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            logger.debug("Adjusted column widths in Excel")

            # Save file
            filename = f"query_report_{uuid.uuid4().hex[:8]}.xlsx"
            wb.save(filename)
            logger.info("Excel report saved: %s", filename)
            return filename
        except Exception as e:
            logger.error("Error creating Excel report: %s", str(e), exc_info=True)
            raise

    async def process_query(self, query):
        """Main function to process user query"""
        logger.info("Starting query processing for: %s", query)
        try:
            # Step 1: Parse query
            parsed_query = await self.parse_query(query)
            
            # Step 2: Generate MCP workflow
            workflow = await self.generate_mcp_workflow(parsed_query)
            
            # Step 3: Scrape data
            all_results = []
            async with async_playwright() as p:
                browser = await p.chromium.launch(headless=True)
                page = await browser.new_page()
                
                # Handle dialogs
                async def handle_dialog(dialog):
                    logger.info("Dialog detected: %s", dialog.message())
                    await dialog.accept()
                page.on("dialog", handle_dialog)
                logger.debug("Dialog handler set up")

                for site_workflow in workflow:
                    site_results = await self.scrape_site(page, site_workflow['site'], site_workflow['steps'])
                    all_results.extend(site_results)
                
                await page.close()
                await browser.close()
                logger.debug("Browser and page closed")
            
            # Step 4: Generate Excel report
            if all_results:
                filename = self.create_excel_report(all_results, query)
                logger.info("Query processing completed successfully")
                return f"Excel report generated: {filename}"
            else:
                logger.warning("No results found for query: %s", query)
                return "No results found for the query."
                
        except Exception as e:
            logger.error("Error processing query '%s': %s", query, str(e), exc_info=True)
            return f"Error processing query: {str(e)}"
        finally:
            # Ensure event loop cleanup
            try:
                loop = asyncio.get_running_loop()
                tasks = [task for task in asyncio.all_tasks(loop) if task is not asyncio.current_task(loop)]
                for task in tasks:
                    task.cancel()
                loop.run_until_complete(loop.shutdown_asyncgens())
            except Exception as e:
                logger.error("Error during event loop cleanup: %s", str(e), exc_info=True)

async def main():
    processor = QueryProcessor(
        llm_provider="ollama",  # Test with Ollama
        ollama_model="llama3"
    )
    query = "Find me laptops under ₹50,000"
    result = await processor.process_query(query)
    print(result)
    logger.info("Main function completed with result: %s", result)

if __name__ == "__main__":
    asyncio.run(main())